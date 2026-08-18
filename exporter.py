"""
Módulo de Exportação de Leads
Exporta os dados formatados para Excel (.xlsx) e JSON com layout profissional.
"""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR


def format_brazilian_phone(raw_phone: str) -> Dict[str, str]:
    """
    Formata o número de telefone no padrão brasileiro e gera o link do WhatsApp.
    """
    if not raw_phone:
        return {"phone_formatted": "", "whatsapp_link": ""}

    # Remove tudo que não for dígito
    digits = re.sub(r'\D', '', str(raw_phone))

    # Remove zero inicial se houver (ex: 011 -> 11)
    if digits.startswith('0') and len(digits) in (11, 12):
        digits = digits[1:]

    # Trata código do país (55)
    if not digits.startswith('55') and len(digits) in (10, 11):
        digits_with_country = f"55{digits}"
    else:
        digits_with_country = digits

    # Formatação visual (XX) XXXXX-XXXX ou (XX) XXXX-XXXX
    formatted = str(raw_phone).strip()
    clean_national = digits[2:] if digits.startswith('55') and len(digits) >= 12 else digits

    if len(clean_national) == 11:
        formatted = f"({clean_national[:2]}) {clean_national[2:7]}-{clean_national[7:]}"
    elif len(clean_national) == 10:
        formatted = f"({clean_national[:2]}) {clean_national[2:6]}-{clean_national[6:]}"

    whatsapp_link = f"https://wa.me/{digits_with_country}" if len(digits_with_country) >= 12 else ""

    return {
        "phone_formatted": formatted,
        "whatsapp_link": whatsapp_link
    }


def export_leads(leads: List[Dict[str, Any]], query_name: str) -> Dict[str, Path]:
    """
    Exporta a lista de leads para arquivos .xlsx e .json.
    Retorna os caminhos dos arquivos gerados.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    clean_query = re.sub(r'[^\w\-_]', '_', query_name).strip('_')
    
    excel_path = OUTPUT_DIR / f"leads_{clean_query}_{timestamp}.xlsx"
    json_path = OUTPUT_DIR / f"leads_{clean_query}_{timestamp}.json"

    # 1. Preparar lista formatada
    processed_leads = []
    for lead in leads:
        phone_info = format_brazilian_phone(lead.get("phone") or lead.get("whatsapp_site") or "")
        wa_link = lead.get("whatsapp_site") or phone_info["whatsapp_link"]
        
        item = {
            "Nome do Estabelecimento": lead.get("name", ""),
            "Nicho / Categoria": lead.get("category", ""),
            "Telefone": phone_info["phone_formatted"],
            "Link WhatsApp": wa_link,
            "E-mail de Contato": lead.get("email") or "",
            "Instagram": lead.get("instagram") or "",
            "Site Oficial": lead.get("website") or "",
            "Endereço": lead.get("address") or "",
            "Nota Google": lead.get("rating") or "",
            "Total de Avaliações": lead.get("reviews_count") or 0,
            "Lead Score": lead.get("lead_score", "3/5"),
            "Qualificação IA": lead.get("ai_qualification", ""),
            "Pitch WhatsApp (Pronto para Enviar)": lead.get("cold_pitch_whatsapp", ""),
            "Link Google Maps": lead.get("maps_url", ""),
        }
        processed_leads.append(item)

    # 2. Exportar JSON
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(leads, f, ensure_ascii=False, indent=2)

    # 3. Exportar Excel com OpenPyXL para visual profissional
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Fitness"
    ws.views.sheetView[0].showGridLines = True

    # Cores e Estilos
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")  # Azul Marinho Escuro
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="0000FF", underline="single")
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    headers = list(processed_leads[0].keys()) if processed_leads else []
    ws.append(headers)

    # Estilizar Cabeçalho
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # Inserir Dados
    for row_idx, lead in enumerate(processed_leads, start=2):
        row_data = list(lead.values())
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 22
        
        is_even = (row_idx % 2 == 0)
        current_fill = zebra_fill if is_even else None

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if current_fill:
                cell.fill = current_fill

            # Alinhamento
            col_name = headers[col_idx - 1]
            if col_name in ["Nota Google", "Total de Avaliações", "Lead Score", "Telefone"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "Pitch WhatsApp (Pronto para Enviar)":
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Formatação de Links clicáveis
            if str(value).startswith("http://") or str(value).startswith("https://"):
                cell.hyperlink = str(value)
                cell.font = link_font

    # Ajuste automático de largura das colunas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        col_name = str(col[0].value)
        
        for cell in col:
            val_str = str(cell.value or "")
            # Limita tamanho para colunas longas como Pitch
            if len(val_str) > max_len:
                max_len = len(val_str)
        
        if col_name == "Pitch WhatsApp (Pronto para Enviar)":
            ws.column_dimensions[col_letter].width = 50
        elif col_name in ["Nome do Estabelecimento", "Endereço", "Qualificação IA"]:
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 25), 45)
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(excel_path)

    return {
        "excel": excel_path,
        "json": json_path
    }
